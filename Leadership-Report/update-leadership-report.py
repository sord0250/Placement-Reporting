# Once called by the 'email-leadership-report' script, this file accesses the excel and proceeds to update all of the tables within using openpyxl. 
# It connects and queries the mySQL database first to get the updated information.
# Then it cycles through the different tables and updates them accordingly. 
# That all sounds pretty straightforward -- it's not. I'll explain it the best I can.

import os
import sys
import re
import datetime as dt
from typing import Dict, List, Tuple
from datetime import date
import mysql.connector
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableColumn
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Border, Side
from pathlib import Path
from dotenv import load_dotenv

# ----------------------------
# 1) Global Variables
# ----------------------------

# Loads in the DATABASE variables so it can connect
BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")
DB_HOST = os.environ["DB_HOST"]
DB_USER = os.environ["DB_USER"]
DB_PASSWORD = os.environ["DB_PASSWORD"]
DB_NAME = os.environ["DB_NAME"]

# Program list (order matters and will be used for summary/program outputs)
PROGRAMS = [
    "BSAcc", "BSEDM", "BSEnt", "BSFin", "BSGSCM", "BSHRM",
    "BSIS", "BSMgt", "BSMktg", "BSStrat", "MAcc", "MBA", "MISM", "MPA",
]

# Gets today's date to use as column names
RUN_DATE = dt.date.today()
RUN_DATE_LABEL = RUN_DATE.strftime("%m/%d/%Y")

# Status labels used for placement calculation (must match SQL result strings exactly)
STATUS_ACCEPTED = "Accepted an offer"
STATUS_SEEKING = "Actively seeking"
STATUS_NOT_REPORTED = "Not Reported"

# Excel Sheet Names
SHEET_SUMMARY_FT = "Summary - Full Time"
SHEET_TOTAL_FT = "Total - Full Time"
SHEET_BYPROG_FT = "By Program - Full Time"
SHEET_TOTAL_INT = "Total - Internships"
SHEET_BYPROG_INT = "By Program - Internships"

# Excel Table Names (except for by program table names)
TABLE_SUMMARY = "summary"
TABLE_TOTAL_FT_MRF = "FT_total_mrf"
TABLE_TOTAL_FT_WH = "FT_total_wh"
TABLE_TOTAL_INT_MRF = "INT_total_mrf"
TABLE_TOTAL_INT_WH = "INT_total_wh"

# Rows that the script knows to avoid, as they use a different calculation for their field
IGNORE_LABELS = {"total", "class size", "% placed", "placement %"}

# ----------------------------
# 1) Special Functions
# ----------------------------

# Checks whether a string is in the IGNORE LABELS
def is_ignored_label(s: str) -> bool:
    if not isinstance(s, str):
        return False
    return s.strip().lower() in IGNORE_LABELS

# Changed labels from Total to CLass Size for clarity. Use it as a safety net now to ensure it doesn't get edited
def relabel_total_row(ws: Worksheet, min_row: int, max_row: int, min_col: int, new_label: str = "Class Size"):
    """Rename the first-column label of the 'total' row to new_label."""
    total_row_idx, _ = find_total_and_placement_rows(ws, min_row, max_row, min_col)
    ws.cell(row=total_row_idx, column=min_col, value=new_label)

# Helpers that creates the By Program table names. Excel doesn't like MBA and MPA apparently so they are done a little different
# Full Time Placement Tables
def byprog_full_names(prog: str) -> Tuple[str, str]:
    """Return (most_recent_table_name, history_table_name) for Full-Time sheet."""
    if prog in ("MBA", "MPA"):
        return f"{prog}_1", f"{prog}_2"
    return f"{prog}1", f"{prog}2"

# Internship Placement Tables
def byprog_int_names(prog: str) -> Tuple[str, str]:
    """Return (most_recent_table_name, history_table_name) for Internships sheet."""
    return f"{prog}_int1", f"{prog}_int2"

# ----------------------------
# 2) SQL
# ----------------------------

# It builds a comma-separated list of SQL parameter placeholders wrapped in parentheses. It'll allow us to query for each program separetly at the same time.
def build_program_in_clause(placeholders: int) -> str:
    return "(" + ",".join(["%s"] * placeholders) + ")"

# Creates the summary sheet: compares each program to each other
SQL_SUMMARY_TEMPLATE = """
SELECT
    program,
    SUM(job_search_status = 'Accepted an offer') AS offer_accepted,
    SUM(job_search_status = 'Actively seeking') AS still_seeking,
    SUM(CASE WHEN COALESCE(job_search_status,'') IN ('Not Reported','No Recent Information Available','') THEN 1 ELSE 0 END) AS no_info,
    SUM(job_search_status LIKE 'Not seeking%') AS not_seeking,
    SUM(CASE WHEN is_international = 1 AND (work_authorization NOT IN ('U.S. Permanent Resident', 'U.S. Citizen') OR work_authorization IS NULL) THEN 1 ELSE 0 END) AS intl_all,
    COUNT(*) AS total
FROM msmdatabase.bcc_student_view
WHERE ((class_of = 2026 and enroll_status IN ("Enrolled", "Graduated")) or (class_of IN (2024, 2025) and enroll_status = "Enrolled"))
  AND program NOT IN ('EMBA','EMPA','StratMnr')
  AND program IN {IN_LIST}
  AND enroll_status IN ('Enrolled','Graduated')
  AND record_status = 'A'
  AND semester_byu NOT IN (20265, 20275, 20285)
GROUP BY program
ORDER BY program;
"""

# Creates the second sheet that has full time MSB class totals split up by job_search_status
SQL_TOTAL_FULL = """
SELECT
    COALESCE(job_search_status, 'Not Reported') AS job_search_status,
    COUNT(*) AS count
FROM msmdatabase.bcc_student_view
WHERE ((class_of = 2026 and enroll_status IN ("Enrolled", "Graduated")) or (class_of IN (2024, 2025) and enroll_status = "Enrolled"))
  AND program NOT IN ('EMBA','EMPA','StratMnr')
  AND enroll_status IN ('Enrolled','Graduated')
  AND record_status = 'A'
  AND semester_byu NOT IN (20265, 20275, 20285)
GROUP BY COALESCE(job_search_status, 'Not Reported')
ORDER BY job_search_status;
"""

# Creates the fourth sheet that has internship MSB class totals split up by job_search_status
SQL_TOTAL_INT = """
SELECT
    COALESCE(internship_search_status, 'Not Reported') AS internship_search_status,
    COUNT(*) AS count
FROM msmdatabase.bcc_student_view
WHERE class_of IN ('2027', '2028', '2029')
  AND program NOT IN ('EMBA','EMPA','StratMnr')
  AND enroll_status IN ('Enrolled','Graduated')
  AND record_status = 'A'
  AND semester_byu NOT IN (20265, 20275, 20285)
GROUP BY COALESCE(internship_search_status, 'Not Reported')
ORDER BY internship_search_status;
"""

# gets each program's full time job search status
SQL_BY_PROGRAM_FULL = """
SELECT
    COALESCE(job_search_status, 'Not Reported') AS job_search_status,
    COUNT(*) AS count
FROM msmdatabase.bcc_student_view
WHERE ((class_of = 2026 and enroll_status IN ("Enrolled", "Graduated")) or (class_of IN (2024, 2025) and enroll_status = "Enrolled"))
  AND program NOT IN ('EMBA','EMPA','StratMnr')
  AND program = %s
  AND enroll_status IN ('Enrolled','Graduated')
  AND record_status = 'A'
  AND semester_byu NOT IN (20265, 20275, 20285)
GROUP BY COALESCE(job_search_status, 'Not Reported')
ORDER BY job_search_status;
"""

# gets each program's internship job search status
SQL_BY_PROGRAM_INT = """
SELECT
    COALESCE(internship_search_status, 'Not Reported') AS internship_search_status,
    COUNT(*) AS count
FROM msmdatabase.bcc_student_view
WHERE class_of IN ('2027', '2028', '2029')
  AND program NOT IN ('EMBA','EMPA','StratMnr')
  AND program = %s
  AND enroll_status IN ('Enrolled','Graduated')
  AND record_status = 'A'
  AND semester_byu NOT IN (20265, 20275, 20285)
GROUP BY COALESCE(internship_search_status, 'Not Reported')
ORDER BY internship_search_status;
"""

# This executes each SQL query using the connection to the DB
def fetch_rows(cursor, sql: str, params: Tuple = ()) -> List[Tuple]:
    cursor.execute(sql, params)
    return list(cursor.fetchall())

# ----------------------------
# 3) Excel Functions 
# ----------------------------

# Helps distinguish between Full Time and Internship tables
def expected_header_for_table(ws: Worksheet, tbl_name: str) -> str:
    title = (ws.title or "").lower()
    name  = (tbl_name or "").lower()
    if ("internship" in title) or ("_int" in name) or name.startswith("int_total"):
        return "Internship Search Status"
    return "Job Search Status"
 
# Finds the header row of the table
def detect_header_row(ws: Worksheet, min_row: int, max_row: int, min_col: int, *, expected_first_header: str) -> int:
    first_col_letter = get_column_letter(min_col)
    exp = expected_first_header.strip().lower()
 
    # 1) Tries the Exact match, which should generally work
    for r in range(min_row, max_row + 1):
        v = ws[f"{first_col_letter}{r}"].value
        if isinstance(v, str) and v.strip().lower() == exp:
            return r
 
    # 2) Loose match (handles minor wording/casing)
    pat = re.compile(r"search\s+status", re.IGNORECASE)
    for r in range(min_row, max_row + 1):
        v = ws[f"{first_col_letter}{r}"].value
        if isinstance(v, str) and pat.search(v):
            return r
 
    # 3) Fallback
    return min_row

# Accesses a table within a worksheet or tab
def get_table(ws: Worksheet, name: str) -> Table:
    if name not in ws.tables:
        raise RuntimeError(f"Expected table '{name}' not found on sheet '{ws.title}'.")
    return ws.tables[name]

# Returns the different table bounds so the size of the table is known
def table_bounds(ref: str) -> Tuple[int, int, int, int]:
    start, end = ref.split(":")
    import re
    cell_re = re.compile(r"([A-Z]+)(\d+)")
    c1 = cell_re.fullmatch(start).groups()
    c2 = cell_re.fullmatch(end).groups()
    min_col = column_index_from_string(c1[0])
    min_row = int(c1[1])
    max_col = column_index_from_string(c2[0])
    max_row = int(c2[1])
    return min_row, max_row, min_col, max_col
 
# Ensures that the metadata of each table matches what actually now exists in the excel.
# Without this, the excel sheet had to be repaired each time it was opened, which was not what we wanted.
def set_table_ref(ws: Worksheet, tbl: Table, min_row: int, max_row: int, min_col: int, max_col: int):
    # 1) Update range + autofilter
    new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    tbl.ref = new_ref
    if getattr(tbl, "autoFilter", None) is not None:
        tbl.autoFilter.ref = new_ref
 
    # 2) Get the mutable list of <tableColumn> elements (handles openpyxl variants)
    tc_container = getattr(tbl, "tableColumns", None)
    tc_list = getattr(tc_container, "tableColumn", None)
    if tc_list is None:
        # some versions store the list directly in tableColumns
        tc_list = tc_container
 
    # 3) Make counts match the width of tbl.ref
    width = max_col - min_col + 1
    meta_count = len(tc_list) if tc_list is not None else 0
 
    # Initialize if somehow missing
    if tc_list is None:
        raise RuntimeError("Table has no tableColumns list; consider upgrading openpyxl.")
 
    # 3a) Trim extras (keep the leftmost N; preserves attributes on existing columns)
    if meta_count > width:
        del tc_list[width:]
        meta_count = width
 
    # 3b) Append missing entries using header cells; keep names unique
    existing_names = {tc.name for tc in tc_list}
    next_id = max((tc.id for tc in tc_list), default=0) + 1
 
    for offset in range(meta_count, width):
        c = min_col + offset  # column index for this new metadata slot
        raw = ws.cell(row=min_row, column=c).value
        base = (str(raw).strip() if raw not in (None, "") else f"Column{offset+1}")
        name, k = base, 1
        while name in existing_names:
            k += 1
            name = f"{base}_{k}"
        existing_names.add(name)
        tc_list.append(TableColumn(id=next_id, name=name))
        next_id += 1
 
    # 4) Lightweight sanity check (no writes to .count — it's derived on save)
    if len(tc_list) != width:
        raise RuntimeError(
            f"Table '{getattr(tbl, 'displayName', '<unnamed>')}' metadata columns={len(tc_list)} "
            f"but width={width} for ref {tbl.ref}"
        )

# Locates the total and placement rows (hardcoded)
def find_total_and_placement_rows(ws: Worksheet, min_row: int, max_row: int, min_col: int) -> Tuple[int, int]:
    total_idx = max_row - 1
    placement_idx = max_row

    return total_idx, placement_idx

# Calculates the placement percentage for the table
def placement_percent(accepted: float, seeking: float, not_reported: float) -> float:
    denom = (accepted or 0) + (seeking or 0) + (not_reported or 0)
    if denom <= 0:
        return 0.0
    return round((accepted or 0) * 100.0 / denom, 2)

# Formats the percents the same way across the board: with two decimals
def write_percent(cell, value: float):
    cell.value = value / 100.0
    cell.number_format = "0.00%"

# Sets the data column's headers to the current date
def ensure_header(ws: Worksheet, header_row: int, label_col: int, data_cols: List[int], header_label: str, force_append: bool=False):
    """
    For MRF: set the single data column header to run date label.
    For WH: add a new column at right with the run date header and return new data_cols list including it.
    """
    if force_append:
        new_col_idx = data_cols[-1] + 1
        ws.cell(row=header_row, column=new_col_idx, value=header_label)
        return data_cols + [new_col_idx]
    
    # MRF (exactly one data column)
    if len(data_cols) == 1:
        ws.cell(row=header_row, column=data_cols[0], value=header_label)
        return data_cols

    # WH (append new column)
    new_col_idx = data_cols[-1] + 1
    ws.cell(row=header_row, column=new_col_idx, value=header_label)

    return data_cols + [new_col_idx]

# Combines most of the functions to update the MRF (Most Recent Friday) tables
def update_mrf_table(
    ws: Worksheet, tbl_name: str, results: List[Tuple[str, int]], status_field: str
):
    """
    Replace counts for the single latest column and set its header to RUN_DATE_LABEL.
    - status_field is 'job_search_status' or 'internship_search_status' (only used for error messages).
    """
    right_align = Alignment(horizontal='right')
    print(tbl_name)
    tbl = get_table(ws, tbl_name)
    min_row, max_row, min_col, max_col = table_bounds(tbl.ref)
    header_expected = expected_header_for_table(ws, tbl_name)
    header_row = detect_header_row(ws, min_row, max_row, min_col, expected_first_header=header_expected)
    label_col = min_col
    data_cols = list(range(min_col + 1, max_col + 1))

    # Expect structure: first column = status label, one numeric data column (MRF)
    if len(data_cols) != 1:
        raise RuntimeError(f"MRF table '{tbl_name}' should have exactly 1 data column; found {len(data_cols)}.")

    # Update header to run date
    ensure_header(ws, header_row, label_col, data_cols, RUN_DATE_LABEL)
    # Updates Metadata so that it matches new Column name
    tc_container = getattr(tbl, "tableColumns", None)
    tc_list = getattr(tc_container, "tableColumn", None) or tc_container
    idx = data_cols[0] - min_col
    tc_list[idx].name = str(ws.cell(row=header_row, column=data_cols[0]).value or f"Column{idx+1}").strip()

    # Map of status->count from SQL
    sql_map: Dict[str, int] = {r[0]: int(r[1]) for r in results}

    # Find Total/Placement rows
    total_row_idx, placement_row_idx = find_total_and_placement_rows(ws, min_row, max_row, min_col)
    relabel_total_row(ws, min_row, max_row, min_col, new_label="Class Size")

    # Walk data rows, write counts; collect which statuses exist
    seen = set()
    for r in range(header_row + 1, max_row + 1):
        label = ws.cell(row=r, column=label_col).value
        if label is None:
            continue
        lstr = str(label).strip()
        if is_ignored_label(lstr):
            continue
        if lstr in sql_map:
            ws.cell(row=r, column=data_cols[0], value=int(sql_map[lstr]))
            seen.add(lstr)
        else:
            ws.cell(row=r, column=data_cols[0], value='-').alignment = right_align
            
    # Recompute totals and placement for the single column
    compute_totals_and_placement(ws, min_row, max_row, min_col, [data_cols[0]])

# Combines all the functions to update the WH tables (Weekly History) tables
def update_wh_table(
    ws: Worksheet, tbl_name: str, results: List[Tuple[str, int]], status_field: str
):
    """
    Append a new column to the right for WH tables, labeled with RUN_DATE_LABEL, and populate counts.
    Insert any new statuses above the Total row; zero-fill older columns for those new rows.
    """
    right_align = Alignment(horizontal='right')
    tbl = get_table(ws, tbl_name)
    min_row, max_row, min_col, max_col = table_bounds(tbl.ref)
    header_expected = expected_header_for_table(ws, tbl_name)
    header_row = detect_header_row(ws, min_row, max_row, min_col, expected_first_header=header_expected)
    label_col = min_col
    existing_data_cols = list(range(min_col + 1, max_col + 1))

    # Add new column header
    new_data_cols = ensure_header(ws, header_row, label_col, existing_data_cols, RUN_DATE_LABEL, force_append=True)
    if len(new_data_cols) == len(existing_data_cols):
        # Should not happen for WH; defensive
        raise RuntimeError(f"WH table '{tbl_name}' did not get a new column appended.")

    newest_col = new_data_cols[-1]
    # Expand the table to include the new column
    set_table_ref(ws, tbl, min_row, max_row, min_col, newest_col)

    # Map SQL results
    sql_map: Dict[str, int] = {r[0]: int(r[1]) for r in results}

    # Find Total/Placement rows
    total_row_idx, placement_row_idx = find_total_and_placement_rows(ws, min_row, max_row, min_col)
    relabel_total_row(ws, min_row, max_row, min_col, new_label="Class Size")

    # Format Line above Total
    newest_col = new_data_cols[-1]
    row_idx = total_row_idx - 1
    thin_border = Border(bottom=Side(style="thin", color="000000"))
    cell = ws.cell(row=row_idx, column=newest_col)
    cell.border = thin_border

    # Fill existing statuses
    seen = set()
    for r in range(header_row + 1, max_row + 1):
        label = ws.cell(row=r, column=label_col).value
        if label is None:
            continue
        lstr = str(label).strip()
        if is_ignored_label(lstr):
            continue
        if lstr in sql_map:
            ws.cell(row=r, column=newest_col, value=int(sql_map[lstr]))
            seen.add(lstr)
        else:
            ws.cell(row=r, column=newest_col, value='-').alignment = right_align

    # Recompute totals and placement for ALL columns (safer)
    compute_totals_and_placement(ws, min_row, max_row, min_col, new_data_cols)

# Totals all of the data to get class size AND creates the placement percentage. These are the special functions that INGORE labels made sure to skip
def compute_totals_and_placement(ws: Worksheet, min_row: int, max_row: int, min_col: int, data_cols: List[int]):
    """
    For each data column in data_cols:
    - Write Total = sum of all numeric rows (excluding 'Total' and '% Placed')
    - Write Placement % = Accepted / (Accepted + Seeking + Not Reported)
    """
    label_col = min_col
    total_row_idx, placement_row_idx = find_total_and_placement_rows(ws, min_row, max_row, min_col)

    # Build a mapping from status label -> row index
    status_to_row: Dict[str, int] = {}
    for r in range(min_row + 1, max_row + 1):
        label = ws.cell(row=r, column=label_col).value
        if not isinstance(label, str):
            continue
        if is_ignored_label(label):
            continue
        status_to_row[label.strip()] = r

    # For each numeric column
    for col in data_cols:
        # Total
        running = 0
        for lbl, row_idx in status_to_row.items():
            val = to_int(ws.cell(row=row_idx, column=col).value)
            try:
                running += int(val or 0)
            except Exception:
                running += 0
        ws.cell(row=total_row_idx, column=col, value=running)

        # Placement %
        acc_val = to_int(ws.cell(row=status_to_row.get(STATUS_ACCEPTED, 0), column=col).value if STATUS_ACCEPTED in status_to_row else 0)
        seek_val = to_int(ws.cell(row=status_to_row.get(STATUS_SEEKING, 0), column=col).value if STATUS_SEEKING in status_to_row else 0)
        nr_val = to_int(ws.cell(row=status_to_row.get(STATUS_NOT_REPORTED, 0), column=col).value if STATUS_NOT_REPORTED in status_to_row else 0)

        pct = placement_percent(acc_val or 0, seek_val or 0, nr_val or 0)
        write_percent(ws.cell(row=placement_row_idx, column=col), pct)

# for counting totals and creating placement percentage: ensure they are ints and no data cells are skipped
def to_int(v):
    try:
        # Handle blanks and "-" as zero
        if v in (None, "", "-"):
            return 0
        # If values might have commas like "1,234", strip them:
        return int(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return 0

# Updates the summary sheet that compares all the programs progress
def update_summary_sheet(ws: Worksheet, rows: List[Tuple]):
    """
    Rows from SQL_SUMMARY: program, offer_accepted, still_seeking, no_info, not_seeking, intl_all, total
    Write into the 'summary' table in PROGRAMS order with the required columns:
      Program | % Placed | Offers Accepted | Still Seeking | Int'l | No Info* | Not Seeking | Total | % NS** | % Null
    """
    tbl = get_table(ws, TABLE_SUMMARY)
    min_row, max_row, min_col, max_col = table_bounds(tbl.ref)
    header_row = min_row

    # Build dict by program
    by_prog: Dict[str, Dict[str, int]] = {}
    for (prog, offer_accepted, still_seeking, no_info, not_seeking, intl_all, total) in rows:
        by_prog[str(prog)] = {
            "offer_accepted": int(offer_accepted or 0),
            "still_seeking": int(still_seeking or 0),
            "no_info": int(no_info or 0),
            "not_seeking": int(not_seeking or 0),
            "intl_all": int(intl_all or 0),
            "total": int(total or 0),
        }

    # Expected 10 columns; we will write values in place. We assume the table already has enough rows.
    # Fail fast if not enough rows to fit all programs.
    needed_rows = len(PROGRAMS)
    available_rows = max_row - header_row
    if available_rows < needed_rows:
        raise RuntimeError(f"Summary table '{TABLE_SUMMARY}' has {available_rows} data rows; needs {needed_rows} for all programs.")

    # Identify column indices by header (to be robust to column ordering)
    headers = {}
    for c in range(min_col, max_col + 1):
        hdr = ws.cell(row=header_row, column=c).value
        if isinstance(hdr, str):
            headers[hdr.strip().lower()] = c

    required_headers = {
        "program", "% placed", "offers accepted", "still seeking", "int'l",
        "no info*", "not seeking", "total", "% ns**", "% null"
    }
    missing_headers = [h for h in required_headers if h not in headers]
    if missing_headers:
        raise RuntimeError(f"Summary headers missing or mismatched: {missing_headers}")

    # Write rows in PROGRAMS order
    r = header_row + 1
    for prog in PROGRAMS:
        data = by_prog.get(prog, {
            "offer_accepted": 0, "still_seeking": 0, "no_info": 0,
            "not_seeking": 0, "intl_all": 0, "total": 0
        })
        # Compute percents
        pct_placed = placement_percent(
            data["offer_accepted"],
            data["still_seeking"],
            data["no_info"]
        )
        pct_ns = round((data["not_seeking"] * 100.0 / data["total"]), 2) if data["total"] else 0.0
        pct_null = round((data["no_info"] * 100.0 / data["total"]), 2) if data["total"] else 0.0

        # Write
        ws.cell(row=r, column=headers["program"], value=prog)
        write_percent(ws.cell(row=r, column=headers["% placed"]), pct_placed)
        ws.cell(row=r, column=headers["offers accepted"], value=data["offer_accepted"])
        ws.cell(row=r, column=headers["still seeking"], value=data["still_seeking"])
        ws.cell(row=r, column=headers["int'l"], value=data["intl_all"])
        ws.cell(row=r, column=headers["no info*"], value=data["no_info"])
        ws.cell(row=r, column=headers["not seeking"], value=data["not_seeking"])
        ws.cell(row=r, column=headers["total"], value=data["total"])
        write_percent(ws.cell(row=r, column=headers["% ns**"]), pct_ns)
        write_percent(ws.cell(row=r, column=headers["% null"]), pct_null)

        r += 1

# ----------------------------
# 5) Main workflow: connect to DB -> run SQL queries -> open Excel workbook -> update each of the sheets -> save and create a copy for history
# ----------------------------

def main():
    template_path = os.path.join(os.path.dirname(__file__), "weekly_placement_report.xlsx")

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found at: {template_path}")

    # Connect DB
    conn = mysql.connector.connect(
        host=DB_HOST, user=DB_USER, password=DB_PASSWORD, database=DB_NAME, autocommit=False
    )
    cur = conn.cursor()

    # Build and run summary SQL with IN clause for PROGRAMS
    in_clause = build_program_in_clause(len(PROGRAMS))
    sql_summary = SQL_SUMMARY_TEMPLATE.format(IN_LIST=in_clause)
    summary_rows = fetch_rows(cur, sql_summary, tuple(PROGRAMS))

    # Totals
    total_ft_rows = fetch_rows(cur, SQL_TOTAL_FULL)
    total_int_rows = fetch_rows(cur, SQL_TOTAL_INT)

    # Per-program buckets
    byprog_ft: Dict[str, List[Tuple[str, int]]] = {}
    byprog_int: Dict[str, List[Tuple[str, int]]] = {}
    for prog in PROGRAMS:
        byprog_ft[prog] = fetch_rows(cur, SQL_BY_PROGRAM_FULL, (prog,))
        byprog_int[prog] = fetch_rows(cur, SQL_BY_PROGRAM_INT, (prog,))

    cur.close()
    conn.close()

    # Open workbook
    wb = load_workbook(template_path, data_only=False)

    # 1) Summary – Full Time
    ws = wb[SHEET_SUMMARY_FT]
    update_summary_sheet(ws, summary_rows)
    print("Updated Summary - Full Time")

    # 2) Total – Full Time (MRF replace & WH append)
    ws_ft_total = wb[SHEET_TOTAL_FT]
    update_mrf_table(ws_ft_total, TABLE_TOTAL_FT_MRF, total_ft_rows, "job_search_status")
    update_wh_table(ws_ft_total, TABLE_TOTAL_FT_WH, total_ft_rows, "job_search_status")
    print("Updated Total - Full Time")

    # 3) By Program – Full Time
    ws_ft_prog = wb[SHEET_BYPROG_FT]
    for prog in PROGRAMS:
        t1, t2 = byprog_full_names(prog)
        update_mrf_table(ws_ft_prog, t1, byprog_ft[prog], "job_search_status")
        update_wh_table(ws_ft_prog, t2, byprog_ft[prog], "job_search_status")
    print("Updated By Program - Full Time")

    # 4) Total – Internships (MRF replace & WH append)
    ws_int_total = wb[SHEET_TOTAL_INT]
    update_mrf_table(ws_int_total, TABLE_TOTAL_INT_MRF, total_int_rows, "internship_search_status")
    update_wh_table(ws_int_total, TABLE_TOTAL_INT_WH, total_int_rows, "internship_search_status")
    print("Updated Total - Internships")

    # 5) By Program – Internships
    ws_int_prog = wb[SHEET_BYPROG_INT]
    for prog in PROGRAMS:
        t1, t2 = byprog_int_names(prog)
        update_mrf_table(ws_int_prog, t1, byprog_int[prog], "internship_search_status")
        update_wh_table(ws_int_prog, t2, byprog_int[prog], "internship_search_status")
    print("Updated By Program - Internships")

    # Save in place (overwrite template as the weekly report, and create the history path)
    wb.save(template_path)
    
    
if __name__ == "__main__":
    try:
        main()
        print(f"Weekly placement report updated successfully: {RUN_DATE_LABEL}")
    except Exception as e:
        # Fail fast with a clear message
        sys.stderr.write(f"[ERROR] {e}\n")
        sys.exit(1)
