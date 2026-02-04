import os
from pathlib import Path
from dotenv import load_dotenv
import mysql.connector
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableColumn
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Border, Side
from datetime import date
import re


# =========================
# 1) Global Variables
# =========================

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")

DB_HOST = os.environ["DB_HOST"]
DB_USER = os.environ["DB_USER"]
DB_PASSWORD = os.environ["DB_PASSWORD"]
DB_NAME = os.environ["DB_NAME"]

# Where the workbook lives
FILEPATH_TEMPLATE = os.getenv("OUTPUT_PATH", str(BASE_DIR / "WeeklyPlacement-{file_label}.xlsx"))

# Run data formatted correctly for column headers
RUN_DATE_LABEL = date.today().strftime("%m/%d/%Y")

# Percent inputs (must match SQL/Excel labels exactly)
STATUS_ACCEPTED = "Accepted an offer"     
STATUS_SEEKING = "Actively seeking"        
STATUS_NOT_REPORTED = "Not Reported"       

# Stylistic Variables
RIGHT_ALIGN = Alignment(horizontal="right")
THIN_BORDER = Border(bottom=Side(style="thin", color="000000"))

# Rows that the script knows to avoid, as they use a different calculation for their field
IGNORE_LABELS = {"total", "class size", "% placed", "placement %"}  


# =========================
# 2) SQL Queries
# =========================

SQL_TOTAL_FULL = """
SELECT
    COALESCE(job_search_status, 'Not Reported') AS job_search_status,
    COUNT(*) AS count
FROM msmdatabase.bcc_student_view
WHERE ((class_of = 2026 and enroll_status IN ("Enrolled", "Graduated")) or (class_of IN (2024, 2025) and enroll_status = "Enrolled"))
  AND program NOT IN ('EMBA','EMPA','StratMnr')
  AND enroll_status IN ('Enrolled','Graduated')
  AND record_status = 'A'
  AND semester_byu NOT IN (20265, 20275, 20285)
GROUP BY COALESCE(job_search_status, 'Not Reported')
ORDER BY job_search_status;
"""

SQL_TOTAL_INT = """
SELECT
    COALESCE(internship_search_status, 'Not Reported') AS internship_search_status,
    COUNT(*) AS count
FROM msmdatabase.bcc_student_view
WHERE class_of IN ('2027', '2028', '2029')
  AND program NOT IN ('EMBA','EMPA','StratMnr')
  AND enroll_status IN ('Enrolled','Graduated')
  AND record_status = 'A'
  AND semester_byu NOT IN (20265, 20275, 20285)
GROUP BY COALESCE(internship_search_status, 'Not Reported')
ORDER BY internship_search_status;
"""

SQL_BY_PROGRAM_FULL = """
SELECT
    COALESCE(job_search_status, 'Not Reported') AS job_search_status,
    COUNT(*) AS count
FROM msmdatabase.bcc_student_view
WHERE ((class_of = 2026 and enroll_status IN ("Enrolled", "Graduated")) or (class_of IN (2024, 2025) and enroll_status = "Enrolled"))
  AND program NOT IN ('EMBA','EMPA','StratMnr')
  AND program = %s
  AND enroll_status IN ('Enrolled','Graduated')
  AND record_status = 'A'
  AND semester_byu NOT IN (20265, 20275, 20285)
GROUP BY COALESCE(job_search_status, 'Not Reported')
ORDER BY job_search_status;
"""

SQL_BY_PROGRAM_INT = """
SELECT
    COALESCE(internship_search_status, 'Not Reported') AS internship_search_status,
    COUNT(*) AS count
FROM msmdatabase.bcc_student_view
WHERE class_of IN ('2027', '2028', '2029')
  AND program NOT IN ('EMBA','EMPA','StratMnr')
  AND program = %s
  AND enroll_status IN ('Enrolled','Graduated')
  AND record_status = 'A'
  AND semester_byu NOT IN (20265, 20275, 20285)
GROUP BY COALESCE(internship_search_status, 'Not Reported')
ORDER BY internship_search_status;
"""

SQL_BSFIN_INT = """
SELECT
    COALESCE(internship_search_status, 'Not Reported') AS internship_search_status,
    COUNT(*) AS count
FROM msmdatabase.bcc_student_view
WHERE class_of = %s
  AND program NOT IN ('EMBA','EMPA','StratMnr')
  AND program = 'BSFin'
  AND enroll_status IN ('Enrolled','Graduated')
  AND record_status = 'A'
  AND semester_byu NOT IN (20265, 20275, 20285)
GROUP BY COALESCE(internship_search_status, 'Not Reported')
ORDER BY internship_search_status;
"""

# =========================
# 3) SMALL Functions
# =========================

# 4 tables per sheet: 1: MRF FT,  2: WH FT,  3: MRF INT,  4: WH INT
# Class sheet uses 'Class1'..'Class4'
# MBA/MPA use underscores because excel doesn't like them; others do not.
# BSFin wants a special internship report ran, so they have two extra tables
def table_names(programs):
    """
    4 tables per sheet:
      1: MRF FT,  2: WH FT,  3: MRF INT,  4: WH INT
    Class sheet uses 'Class1'..'Class4'
    MBA/MPA use underscores; others do not.
    """
    tbl_nms = {"Class": ("Class1", "Class2", "Class3", "Class4")}
    for program in programs:
        if program in ("MPA", "MBA"):
            tbl_nms[program] = (f"{program}_1", f"{program}_2", f"{program}_3", f"{program}_4")
        elif program == "BSFin":
            tbl_nms[program] = (f"{program}1", f"{program}2", f"{program}3", f"{program}4", f"{program}5", f"{program}6")
        else:
            tbl_nms[program] = (f"{program}1", f"{program}2", f"{program}3", f"{program}4")
    return tbl_nms

# Turns a program title (i.e BSacc) into its corresponding file name
def program_to_filename(programs):
    if not programs:
        raise RuntimeError("No programs provided.")
    return programs[0] if len(programs) == 1 else "-".join(programs)

# Executes the SQL queries
def fetch_rows(cur, sql, params=()):
    cur.execute(sql, params)
    return list(cur.fetchall())

# Connects to a specific table on a worksheet in excel
def get_table(ws: Worksheet, name: str) -> Table:
    if name not in ws.tables:
        raise RuntimeError(f"Expected table '{name}' not found on sheet '{ws.title}'.")
    return ws.tables[name]

# Gets the specific table bounds so they can be used to input data
def table_bounds(ref: str):
    start, end = ref.split(":")
    cell_re = re.compile(r"([A-Z]+)(\d+)")
    c1 = cell_re.fullmatch(start).groups()
    c2 = cell_re.fullmatch(end).groups()
    min_col = column_index_from_string(c1[0]); min_row = int(c1[1])
    max_col = column_index_from_string(c2[0]); max_row = int(c2[1])
    return min_row, max_row, min_col, max_col

# Decides whether it is a Full-Time or Internship table
def expected_header_for_table(ws: Worksheet, tbl_name: str) -> str:
    title = (ws.title or "").lower()
    name  = (tbl_name or "").lower()
    if ("internship" in title) or ("_int" in name) or name.endswith("int1") or name.endswith("int2"):
        return "Internship Search Status"  
    return "Job Search Status"              

# Find header row inside the table by header text; fall back to min_row.
def detect_header_row(ws: Worksheet, min_row: int, max_row: int, min_col: int, expected_first_header: str) -> int:
    first_col_letter = get_column_letter(min_col)
    exp = expected_first_header.strip().lower()

    # exact match
    for r in range(min_row, max_row + 1):
        v = ws[f"{first_col_letter}{r}"].value
        if isinstance(v, str) and v.strip().lower() == exp:
            return r

    # loose match: "*search status*"
    pat = re.compile(r"search\s+status", re.IGNORECASE)
    for r in range(min_row, max_row + 1):
        v = ws[f"{first_col_letter}{r}"].value
        if isinstance(v, str) and pat.search(v):
            return r

    return min_row  #fallback

# For MRF: set the single data column header to run date.
# For WH: append a rightmost column with run date if force_append=True.
def ensure_header(ws: Worksheet, header_row: int, data_cols, header_label: str, force_append: bool=False):
    if force_append:
        new_col_idx = data_cols[-1] + 1
        ws.cell(row=header_row, column=new_col_idx, value=header_label)
        return data_cols + [new_col_idx]

    if len(data_cols) == 1:
        ws.cell(row=header_row, column=data_cols[0], value=header_label)
        return data_cols

    new_col_idx = data_cols[-1] + 1
    ws.cell(row=header_row, column=new_col_idx, value=header_label)
    return data_cols + [new_col_idx]

# Ensures that the metadata of each table matches what actually now exists in the excel.
# Without this, the excel sheet had to be repaired each time it was opened, which was not what we wanted.
def set_table_ref(ws: Worksheet, tbl: Table, min_row: int, max_row: int, min_col: int, max_col: int):

    new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    tbl.ref = new_ref
    if getattr(tbl, "autoFilter", None) is not None:
        tbl.autoFilter.ref = new_ref

    tc_container = getattr(tbl, "tableColumns", None)
    tc_list = getattr(tc_container, "tableColumn", None)
    if tc_list is None:
        tc_list = tc_container
    if tc_list is None:
        raise RuntimeError("Table has no tableColumns list; cannot adjust metadata.")

    width = max_col - min_col + 1
    meta_count = len(tc_list)

    # Trim extras (rare)
    if meta_count > width:
        del tc_list[width:]
        meta_count = width

    existing_names = {tc.name for tc in tc_list}
    next_id = max((tc.id for tc in tc_list), default=0) + 1
    for offset in range(meta_count, width):
        c = min_col + offset
        raw = ws.cell(row=min_row, column=c).value
        base = (str(raw).strip() if raw not in (None, "") else f"Column{offset+1}")
        name, k = base, 1
        while name in existing_names:
            k += 1
            name = f"{base}_{k}"
        existing_names.add(name)
        tc_list.append(TableColumn(id=next_id, name=name))
        next_id += 1

    if len(tc_list) != width:
        raise RuntimeError(f"Table metadata columns={len(tc_list)} but width={width} for {tbl.ref}")

def relabel_total_row_to_class_size(ws: Worksheet, label_col, min_row, max_row):
    """Rename 'Total' → 'Class Size' if present."""
    for r in range(min_row, max_row + 1):
        v = ws.cell(row=r, column=label_col).value
        if isinstance(v, str) and v.strip().lower() == "total":
            ws.cell(row=r, column=label_col, value="Class Size")
            return r
    # if already 'Class Size', return that row
    for r in range(min_row, max_row + 1):
        v = ws.cell(row=r, column=label_col).value
        if isinstance(v, str) and v.strip() == "Class Size":
            return r
    # else, assume penultimate row is total
    return max_row - 1

def find_percent_row(ws: Worksheet, label_col, min_row, max_row):
    for r in range(min_row, max_row + 1):
        v = ws.cell(row=r, column=label_col).value
        if isinstance(v, str) and v.strip() == "% Placed":
            return r
    # else, assume last row
    return max_row

def write_dash(cell):
    cell.value = "-"
    cell.alignment = RIGHT_ALIGN

def to_int(v):
    try:
        if v in (None, "", "-"):
            return 0
        return int(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return 0

def placement_percent(accepted: int, seeking: int, not_reported: int) -> float:
    denom = (accepted or 0) + (seeking or 0) + (not_reported or 0)
    if denom <= 0:
        return 0.0
    return round((accepted or 0) * 100.0 / denom, 2)

def write_percent(cell, pct: float):
    cell.value = pct / 100.0
    cell.number_format = "0.00%"

# =========================
# TABLE UPDATERS (MRF/WH)
# =========================

def update_mrf_table(ws: Worksheet, tbl_name: str, sql_rows):
    """
    MRF: two columns total (Status | value). Overwrite the single data column with RUN_DATE_LABEL,
    fill values, then recompute Class Size and % Placed for that column.
    """
    tbl = get_table(ws, tbl_name)
    min_row, max_row, min_col, max_col = table_bounds(tbl.ref)

    header_expected = expected_header_for_table(ws, tbl_name)
    header_row = detect_header_row(ws, min_row, max_row, min_col, header_expected)

    label_col = min_col
    data_cols = list(range(min_col + 1, max_col + 1))
    if len(data_cols) != 1:
        raise RuntimeError(f"MRF table '{tbl_name}' should have exactly 1 data column; found {len(data_cols)}.")

    # header
    ensure_header(ws, header_row, data_cols, RUN_DATE_LABEL)
    # sync tableColumns name for that data column (keeps metadata tidy)
    tc_list = getattr(getattr(tbl, "tableColumns", None), "tableColumn", None) or getattr(tbl, "tableColumns", None)
    if tc_list:
        idx = data_cols[0] - min_col
        tc_list[idx].name = str(ws.cell(row=header_row, column=data_cols[0]).value or f"Column{idx+1}").strip()

    # map SQL to dict
    sql_map = {str(r[0]).strip(): int(r[1]) for r in sql_rows}

    # fill
    for r in range(header_row + 1, max_row + 1):
        label = ws.cell(row=r, column=label_col).value
        if not label:
            continue
        s = str(label).strip()
        if s.lower() in IGNORE_LABELS:
            continue
        if s in sql_map:
            ws.cell(row=r, column=data_cols[0], value=int(sql_map[s]))
        else:
            write_dash(ws.cell(row=r, column=data_cols[0]))

    # totals + % placed for this column
    compute_totals_and_percent(ws, min_row, max_row, min_col, data_cols[-1])

def update_wh_table(ws: Worksheet, tbl_name: str, sql_rows):
    """
    WH: append a new column at the right, label it RUN_DATE_LABEL, fill,
    draw a thin line above Class Size, and compute totals/% placed for that new column.
    """
    tbl = get_table(ws, tbl_name)
    min_row, max_row, min_col, max_col = table_bounds(tbl.ref)

    header_expected = expected_header_for_table(ws, tbl_name)
    header_row = detect_header_row(ws, min_row, max_row, min_col, header_expected)

    label_col = min_col
    existing_data_cols = list(range(min_col + 1, max_col + 1))

    # add header at right
    new_cols = ensure_header(ws, header_row, existing_data_cols, RUN_DATE_LABEL, force_append=True)
    newest_col = new_cols[-1]

    # widen the table safely (preserves metadata)
    set_table_ref(ws, tbl, min_row, max_row, min_col, newest_col)

    # map results
    sql_map = {str(r[0]).strip(): int(r[1]) for r in sql_rows}

    # fill
    for r in range(header_row + 1, max_row + 1):
        label = ws.cell(row=r, column=label_col).value
        if not label:
            continue
        s = str(label).strip()
        if s.lower() in IGNORE_LABELS:
            continue
        if s in sql_map:
            ws.cell(row=r, column=newest_col, value=int(sql_map[s]))
        else:
            write_dash(ws.cell(row=r, column=newest_col))

    # thin border above Class Size for visual separation
    total_row = relabel_total_row_to_class_size(ws, label_col, min_row, max_row)
    ws.cell(row=total_row - 1, column=newest_col).border = THIN_BORDER

    # totals + % placed for newest column
    compute_totals_and_percent(ws, min_row, max_row, min_col, newest_col)

def compute_totals_and_percent(ws: Worksheet, min_row: int, max_row: int, min_col: int, col: int):
    """
    Total = sum of numeric rows (exclude 'Class Size' and '% Placed')
    % Placed = Accepted an offer / (Accepted an offer + Actively seeking + Not Reported)
    """
    label_col = min_col

    # locate special rows
    total_row = relabel_total_row_to_class_size(ws, label_col, min_row, max_row)
    pct_row   = find_percent_row(ws, label_col, min_row, max_row)

    # collect status rows
    status_to_row = {}
    for r in range(min_row + 1, max_row + 1):
        label = ws.cell(row=r, column=label_col).value
        if not isinstance(label, str):
            continue
        s = label.strip()
        if s.lower() in IGNORE_LABELS:
            continue
        status_to_row[s] = r

    # total
    total = 0
    for s, rr in status_to_row.items():
        total += to_int(ws.cell(row=rr, column=col).value)
    ws.cell(row=total_row, column=col, value=total)

    # percent placed
    acc = to_int(ws.cell(row=status_to_row.get(STATUS_ACCEPTED, 0), column=col).value if STATUS_ACCEPTED in status_to_row else 0)
    seek = to_int(ws.cell(row=status_to_row.get(STATUS_SEEKING, 0), column=col).value if STATUS_SEEKING in status_to_row else 0)
    nr  = to_int(ws.cell(row=status_to_row.get(STATUS_NOT_REPORTED, 0), column=col).value if STATUS_NOT_REPORTED in status_to_row else 0)

    pct = placement_percent(acc, seek, nr)
    write_percent(ws.cell(row=pct_row, column=col), pct)

# =========================
# SHEET UPDATER
# =========================

def update_sheet_with_ft_int(ws: Worksheet, table_tuple, ft_rows, int_rows):
    """
    Update 4 tables on a sheet:
      0: MRF FT (ft_rows)
      1: WH  FT (ft_rows)
      2: MRF INT (int_rows)
      3: WH  INT (int_rows)
    """
    t1, t2, t3, t4 = table_tuple
    update_mrf_table(ws, t1, ft_rows)
    update_wh_table(ws, t2, ft_rows)
    update_mrf_table(ws, t3, int_rows)
    update_wh_table(ws, t4, int_rows)

def update_bsfin_with_ft_int(ws: Worksheet, table_tuple, ft_rows, int_2027_rows, int_2028_rows):
    """
    Update 4 tables on a sheet:
      0: MRF FT (ft_rows)
      1: WH  FT (ft_rows)
      2: MRF INT (int_rows)
      3: WH  INT (int_rows)
    """
    t1, t2, t3, t4, t5, t6 = table_tuple
    update_mrf_table(ws, t1, ft_rows)
    update_wh_table(ws, t2, ft_rows)
    update_mrf_table(ws, t3, int_2027_rows)
    update_wh_table(ws, t4, int_2027_rows)
    update_mrf_table(ws, t5, int_2028_rows)
    update_wh_table(ws, t6, int_2028_rows)

# =========================
# MAIN
# =========================

def main(programs):
    # DB
    conn = mysql.connector.connect(
        host=DB_HOST, user=DB_USER, password=DB_PASSWORD, database=DB_NAME, autocommit=False
    )
    cur = conn.cursor()

    # totals
    total_ft = fetch_rows(cur, SQL_TOTAL_FULL)
    total_int = fetch_rows(cur, SQL_TOTAL_INT)

    # per-program
    byProg_ft = {p: fetch_rows(cur, SQL_BY_PROGRAM_FULL, (p,)) for p in programs}
    byProg_int = {p: fetch_rows(cur, SQL_BY_PROGRAM_INT, (p,)) for p in programs}
    BSFin_int_2027 = fetch_rows(cur, SQL_BSFIN_INT, ("2027",))
    BSFin_int_2028 = fetch_rows(cur, SQL_BSFIN_INT, ("2028",))

    cur.close()
    conn.close()

    # workbook
    fileLbl = program_to_filename(programs)
    wb_path = FILEPATH_TEMPLATE.format(file_label=fileLbl)
    wb = load_workbook(wb_path, data_only=False)

    # tables
    tbls = table_names(programs)

    # totals sheet (exact name confirmed earlier)
    class_ws_name = "2026 MSB Overall"
    if class_ws_name not in wb.sheetnames:
        raise RuntimeError(f"Expected sheet '{class_ws_name}' not found.")
    class_ws = wb[class_ws_name]
    update_sheet_with_ft_int(class_ws, tbls["Class"], total_ft, total_int)

    # program sheets
    for program in programs:
        if program not in wb.sheetnames:
            raise RuntimeError(f"Expected program sheet '{program}' not found.")
        elif program == "BSFin":
            ws = wb[program]
            update_bsfin_with_ft_int(ws, tbls[program], byProg_ft[program], BSFin_int_2027, BSFin_int_2028)
        else:
            ws = wb[program]
            update_sheet_with_ft_int(ws, tbls[program], byProg_ft[program], byProg_int[program])

    wb.save(wb_path)
    print(f"Updated: {wb_path}")

if __name__ == "__main__":
    # example
    main(["BSFin"])
